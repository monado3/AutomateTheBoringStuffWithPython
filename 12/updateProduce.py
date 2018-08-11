import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27,
                 }

for row_idx in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row_idx, 1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row_idx, 2).value = PRICE_UPDATES[produce_name]

wb.save('updatedProduceSales.xlsx')
