import pprint

import openpyxl

print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
countryData = {}

print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    state = sheet.cell(row, 2).value
    country = sheet.cell(row, 3).value
    pop = sheet.cell(row, 4).value

    countryData.setdefault(state, {})
    countryData[state].setdefault(country, {'tracts': 0, 'pop': 0})

    countryData[state][country]['tracts'] += 1
    countryData[state][country]['pop'] += int(pop)

print('Writing results...')
with open('census2010.py', 'w') as f:
    f.write('allData = ' + pprint.pformat(countryData))
print('Done.')
